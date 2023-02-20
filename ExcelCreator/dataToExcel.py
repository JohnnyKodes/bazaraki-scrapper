import openpyxl
import os
from Data.carData import CarData
from Data.Constants import COLUMN_TITLES

def setHeaders(worksheet):
    for columnNum, title in enumerate(COLUMN_TITLES, start=1):
        worksheet.cell(row = 1, column = columnNum, value=title)
        
def setCarData(carData, worksheet, brand):
    for rowNum, rowData in enumerate(carData[brand], start=2):
        for columnNum, cellValue in enumerate(rowData.returnIterable(), start=1):
            worksheet.cell(row=rowNum, column= columnNum, value = cellValue) 

def serializeToExcelSheet(carData : CarData):
    workbook = openpyxl.Workbook()
    
    for brand in carData:
        worksheet = workbook.create_sheet(title = brand)
        
        setHeaders(worksheet)
        setCarData(carData, worksheet, brand)    

    path = os.path.join(os.getcwd(), "ScrappedData")

    if not os.path.exists(path):
        os.mkdir(path)

    workbook.save('ScrappedData\BazarakiScrappedData.xlsx')